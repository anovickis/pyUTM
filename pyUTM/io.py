#!/usr/bin/env python
#
# License: BSD 2-clause
# Last Change: Mon Dec 14, 2020 at 04:00 PM +0100

import openpyxl
import re
import yaml

from pyparsing import nestedExpr
from collections import defaultdict
from itertools import zip_longest
from multipledispatch import dispatch
from pathlib import Path
from types import FunctionType
from openpyxl.worksheet.table import Table, TableStyleInfo

from .datatype import range, ColNum, ExcelCell
from .datatype import NetNode
from .common import flatten
from .legacy import PADDING


##############################
# General write/read to file #
##############################

def write_to_file(filename, data, mode='w', eol='\n'):
    with open(filename, mode) as f:
        for row in data:
            f.write(row + eol)


class ReaderWriter(object):
    def __init__(self, filename):
        self.filename = filename


##################
# For CSV output #
##################

def csv_line(node, prop):
    s = ''
    netname = prop['NETNAME']
    attr = prop['ATTR']

    if netname is None:
        s += attr
    elif attr is not None:
        net_head, net_tail = netname.split('_', 1)
        s += (net_head + attr + net_tail)
    else:
        s += netname
    s += ','

    # This should be fine as long as 'node' is a list-like structure.
    for item in node:
        if item is not None:
            s += item
        s += ','

    # Remove the trailing ','
    return s[:-1]


@dispatch((str, Path), dict, FunctionType)
def write_to_csv(filename, data, formatter, **kwargs):
    output = [formatter(k, v) for k, v in data.items()]
    write_to_file(filename, output, **kwargs)


@dispatch((str, Path), list, dict)
def write_to_csv(filename, data, headers, **kwargs):
    header_row = [','.join(headers.keys())]
    body = [','.join([str(entry[k]) for _, k in headers.items()])
            for entry in data]
    write_to_file(filename, header_row+body, **kwargs)


@dispatch((str, Path), list, list)
def write_to_csv(filename, data, headers, **kwargs):
    header_row = [','.join(headers)]
    body = [','.join(map(str, row)) for row in data]
    write_to_file(filename, header_row+body, **kwargs)


#######################
# For Excel documents #
#######################

def parse_cell_range(s, add_one_to_trailing_cell=True):
    initial, final = s.split(':')
    initial_col, initial_row = filter(None, re.split(r'(\d+)', initial))
    final_col, final_row = filter(None, re.split(r'(\d+)', final))

    if add_one_to_trailing_cell:
        return (ColNum(initial_col.upper()), int(initial_row),
                ColNum(final_col.upper())+1, int(final_row)+1)
    else:
        return (ColNum(initial_col.upper()), int(initial_row),
                ColNum(final_col.upper()), int(final_row))


def prepare_descr_for_xlsx_output(descr):
    output = {}
    for k, v in descr.items():
        headers = list(v[0].keys())
        output[k] = []
        output[k].append(headers)

        for i in v:
            temp = [PADDING(i[k]) if k in ['SEAM pin', 'Pigtail pin']
                    else i[k] for k in headers]
            final = [i if i is not None else '' for i in temp]
            output[k].append(final)

    return output


class XLReader(ReaderWriter):
    def read(self, sheets, cell_range, sortby=None, headers=None):
        self.sheets = sheets
        self.cell_range = cell_range
        self.initial_col, self.initial_row, self.final_col, self.final_row = \
            parse_cell_range(cell_range)

        result = []
        # NOTE: The ResourcesWarning is probably due to a lack of encoding in
        # the OS. Ignore it for now.
        wb = openpyxl.load_workbook(self.filename, read_only=True)
        for s in self.sheets:
            ws = wb[str(s)]
            result.append(self.readsheet(ws, sortby=sortby, headers=headers))
        wb.close()
        return result

    def readsheet(self, ws, sortby, headers):
        # We read the full rectangular region to build up a cache. Otherwise if
        # we read one cell at a time, all cells prior to that cell must be read,
        # rendering that method VERY inefficient.
        sheet_region = ws[self.cell_range]
        sheet = dict()
        for row in range(self.initial_row, self.final_row):
            for col in range(self.initial_col, self.final_col):
                sheet[str(col)+str(row)] = \
                    sheet_region[row-self.initial_row][col-self.initial_col]

        if headers is not None:
            data = self.get_data_header_supplied(sheet, headers)
        else:
            data = self.get_data_header_not_supplied(sheet)

        if sortby is not None:
            return sorted(data, key=sortby)
        else:
            return data

    def get_data_header_not_supplied(self, sheet):
        # Read the first row as headers, determine non-empty headers;
        # for all subsequent rows, skip columns without a header.
        headers_non_empty_col = dict()
        for col in range(self.initial_col, self.final_col):
            anchor = str(col) + str(self.initial_row)
            header = sheet[anchor].value
            if header is not None:
                # Note: some of the title contain '\n'. We replace it with an
                # space.
                headers_non_empty_col[str(col)] = header.replace('\n', ' ')

        return self.get_data_header_supplied(sheet, headers_non_empty_col,
                                             initial_row_bump=1)

    def get_data_header_supplied(self, sheet, headers, initial_row_bump=0):
        data = []
        for row in range(self.initial_row+initial_row_bump, self.final_row):
            pin_spec = dict()

            for col in headers.keys():
                anchor = str(col) + str(row)

                name = sheet[anchor].value
                if name is None:
                    pin_spec[headers[col]] = None
                else:
                    try:
                        font_color = sheet[anchor].font.color
                    except AttributeError:
                        font_color = None
                    pin_spec[headers[col]] = ExcelCell(name, font_color)

            data.append(pin_spec)
        return data


class XLWriter(ReaderWriter):
    def write(self, data, **kwargs):
        wb = openpyxl.Workbook()

        # Remove the default sheet
        wb.remove(wb['Sheet'])

        for sheet_name, sheet_data in data.items():
            ws = self.create_sheet(wb, sheet_name)

            headers = sheet_data[0]
            body = sheet_data[1:]
            self.write_table(ws, sheet_name, headers, body, **kwargs)

        wb.save(self.filename)

    @classmethod
    def write_table(cls, ws, table_title, headers, body,
                    initial_row=1, initial_col=ColNum('A'),
                    table_style=TableStyleInfo(
                        name='TableStyleMedium2', showFirstColumn=False,
                        showLastColumn=False, showRowStripes=True,
                        showColumnStripes=False
                    )
                    ):
        raw_data = [headers] + body
        arranged_data, cell_range = cls.rearrange_table(
            raw_data, initial_row, initial_col)

        for row in arranged_data:
            ws.append(row)

        tab = Table(displayName=table_title, ref=cell_range)
        tab.tableStyleInfo = table_style
        ws.add_table(tab)

    @staticmethod
    def create_sheet(wb, name):
        ws = wb.create_sheet(name)
        ws.title = name
        return ws

    @staticmethod
    def rearrange_table(raw, initial_row, initial_col):
        arranged = []

        for r in range(1, initial_row):
            arranged.append([''])

        for r in raw:
            row = []
            for i in range(ColNum('A'), initial_col):
                row.append('')
            for entry in r:
                row.append(entry)
            arranged.append(row)

        # Figure out the cell range of the table
        # Assuming the dimension of each row is the same
        anchor_col = initial_col + len(raw[0]) - 1
        anchor_row = initial_row + len(raw) - 1
        cell_range = '{0}{1}:{2}{3}'.format(
            initial_col, initial_row,
            anchor_col, anchor_row
        )

        return (arranged, cell_range)


####################
# For Pcad netlist #
####################

class NestedListReader(ReaderWriter):
    def read(self):
        return nestedExpr().parseFile(self.filename).asList()[0]


class PcadNaiveReader(NestedListReader):
    # Heavily-modified Zishuo's implementation.
    def read(self, component_postprocessor=lambda x: x.upper()):
        nets = super().read()
        all_nets = {}

        for net in filter(lambda i: isinstance(i, list) and i[0] == 'net',
                          nets):
            netname = net[1].strip('\"')
            all_nets[netname] = []

            for node in \
                    filter(lambda i: isinstance(i, list) and i[0] == 'node',
                           net):

                component, pin = map(
                    component_postprocessor,
                    map(lambda x: x.strip('\"'), node[1:3]))
                all_nets[netname].append((component, pin))

        return all_nets


class PcadReader(PcadNaiveReader):
    def read(self, nethopper, **kwargs):
        all_nets = super().read(**kwargs)
        equivalent_nets = nethopper.do(all_nets)

        self.make_equivalent_nets_identical(all_nets, equivalent_nets)

        return all_nets

    @staticmethod
    def make_equivalent_nets_identical(nets, equivalency):
        for g in equivalency:
            head, tail = g[0], g[1:]
            # Update the head net
            for n in tail:
                nets[head] += nets[n]
                # Now make sure all nets in tail are equivalent to head
                nets[n] = nets[head]


################
# For wirelist #
################

class WirelistNaiveReader(ReaderWriter):
    def read(self, wire_list_name='Wire List'):
        with open(self.filename, 'r') as f:
            raw = f.readlines()
        trees = self.parse_into_trees(raw)
        return self.parse_wire_list(trees[wire_list_name])

    @staticmethod
    def parse_into_trees(lines):
        output = defaultdict(list)
        output['Unnamed']
        output_ptr = output['Unnamed']

        for line in lines:
            if not line.startswith('<<<'):
                output_ptr.append(line.strip())
            else:
                key = line.replace('<<<', '').replace('>>>', '').strip()
                output_ptr = output[key]

        return output

    @staticmethod
    def parse_wire_list(raw):
        output = {}
        garbage = []
        output_ptr = garbage

        for line in raw:
            if not line.startswith('['):
                fields = line.strip().split()
                try:
                    output_ptr.append((fields[0], fields[1]))
                except Exception:
                    pass

            else:
                key = line.replace('[', '').replace(']', '').strip()
                key = ' '.join(key.split(' ')[1:])
                output[key] = []
                output_ptr = output[key]

        return output


############
# For YAML #
############

class YamlReader(ReaderWriter):
    def read(self, flattener=flatten, sortby=None):
        with open(self.filename) as f:
            raw = yaml.safe_load(f)

        for k in raw.keys():
            raw[k] = flattener(raw[k])
            if sortby is not None:
                raw[k] = sorted(raw[k], key=sortby)

        return raw


###############
# For NetNode #
###############

def netnode_to_netlist(nodes):
    nets = defaultdict(list)

    for n in nodes.keys():
        prop = nodes[n]
        netname, attr = prop['NETNAME'], prop['ATTR']

        if netname is None:
            real_netname = attr
        elif attr is not None:
            net_head, net_tail = netname.split('_', 1)
            real_netname = (net_head + attr + net_tail)
        else:
            real_netname = netname

        if n.DCB is not None:
            nets[real_netname].append((n.DCB, n.DCB_PIN))

        if n.PT is not None:
            nets[real_netname].append((n.PT, n.PT_PIN))

    return nets


class NetNodeGen(object):
    def do(self, nets):
        return self.parse_netlist_dict(nets)

    def parse_netlist_dict(self, all_nets_dict):
        net_nodes_dict = {}

        for netname in all_nets_dict.keys():
            net = all_nets_dict[netname]

            dcb_nodes = self.find_node_match_regex(net, re.compile(r'^JD\d+'))
            pt_nodes = self.find_node_match_regex(net, re.compile(r'^JP\d+'))
            other_nodes = list(
                set(net) - set(dcb_nodes) - set(pt_nodes)
            )

            # First, handle DCB-PT connections
            if dcb_nodes and pt_nodes:
                for d, p in zip_longest(dcb_nodes, pt_nodes):
                    net_nodes_dict[self.net_node_gen(d, p)] = {
                        'NETNAME': netname,
                        'ATTR': None
                    }

            # Now if we do have other components...
            if other_nodes and dcb_nodes:
                for d in dcb_nodes:
                    net_nodes_dict[self.net_node_gen(d, None)] = {
                        'NETNAME': netname,
                        'ATTR': None
                    }

            if other_nodes and pt_nodes:
                for p in pt_nodes:
                    net_nodes_dict[self.net_node_gen(None, p)] = {
                        'NETNAME': netname,
                        'ATTR': None
                    }

        return net_nodes_dict

    @staticmethod
    def net_node_gen(dcb_spec, pt_spec, datatype=NetNode):
        try:
            dcb, dcb_pin = dcb_spec
        except Exception:
            dcb = dcb_pin = None

        try:
            pt, pt_pin = pt_spec
        except Exception:
            pt = pt_pin = None

        return datatype(dcb, dcb_pin, pt, pt_pin)

    @staticmethod
    def find_node_match_regex(nodes_list, regex):
        return list(filter(lambda x: regex.search(x[0]), nodes_list))
